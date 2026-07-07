"""
Export generated NEET questions to Excel format.
Groups questions by type, with one sheet per question_type + difficulty.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Short names for sheet titles
TYPE_SHORT = {
    "MCQ": "MCQ",
    "ASSERTION_REASON": "A_R",
    "MATCH_THE_COLUMN": "MTC",
}

HEADERS = [
    "Question",
    "Option A", "Option B", "Option C", "Option D",
    "Accuracy", "Comment",
    "Time to Run",
    "Input Tokens", "Output Tokens", "Total Tokens",
    "Input Cost (₹)", "Output Cost (₹)", "Total Cost (₹)",
    "level change", "reason",
]

# Column widths (index matches HEADERS order, 1-based)
COL_WIDTHS = {
    1: 70,   # Question
    2: 35,   # Option A
    3: 35,   # Option B
    4: 35,   # Option C
    5: 35,   # Option D
    6: 12,   # Accuracy
    7: 20,   # Comment
    8: 14,   # Time to Run
    9: 14,   # Input Tokens
    10: 14,  # Output Tokens
    11: 14,  # Total Tokens
    12: 14,  # Input Cost (₹)
    13: 14,  # Output Cost (₹)
    14: 14,  # Total Cost (₹)
    15: 14,  # level change
    16: 25,  # reason
}


def export_questions_to_excel(
    result: dict,
    time_taken: float = None,
    input_tokens: int = None,
    output_tokens: int = None,
    total_tokens: int = None,
    input_cost: float = None,
    output_cost: float = None,
    total_cost: float = None,
) -> bytes:
    """
    Export questions to an Excel workbook.

    Each question_type gets its own sheet named {TYPE}_{DIFFICULTY}.
    Returns the workbook as bytes (ready for st.download_button or file write).
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    questions = result.get("questions", [])
    metadata = result.get("test_metadata", {})
    difficulty = metadata.get("difficulty", "unknown").upper()

    # Group questions by type
    grouped: dict[str, list] = {}
    for q in questions:
        q_type = q.get("question_type", "MCQ")
        grouped.setdefault(q_type, []).append(q)

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for q_type, q_list in grouped.items():
        short = TYPE_SHORT.get(q_type, q_type.replace(" ", "_").upper())
        sheet_name = f"{short}_{difficulty}"[:31]  # Excel 31-char limit
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Column widths
        from openpyxl.utils import get_column_letter
        for col, width in COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Data rows
        for i, q in enumerate(q_list, start=2):
            q_text = q.get("question_text", "")
            options = q.get("options", {})

            # AR options are fixed — fill if LLM omitted them
            if q_type == "ASSERTION_REASON" and not options:
                options = {
                    "a": "Both Assertion and Reason are true and Reason is the correct explanation of Assertion",
                    "b": "Both Assertion and Reason are true but Reason is NOT the correct explanation of Assertion",
                    "c": "Assertion is true but Reason is false",
                    "d": "Assertion is false but Reason is true",
                }

            row_data = [
                q_text,
                options.get("a", ""),
                options.get("b", ""),
                options.get("c", ""),
                options.get("d", ""),
                "",  # Accuracy (manual)
                "",  # Comment (manual)
                f"{time_taken:.1f}s" if (time_taken is not None and i == 2) else "",
                input_tokens if (input_tokens is not None and i == 2) else "",
                output_tokens if (output_tokens is not None and i == 2) else "",
                total_tokens if (total_tokens is not None and i == 2) else "",
                input_cost if (input_cost is not None and i == 2) else "",
                output_cost if (output_cost is not None and i == 2) else "",
                total_cost if (total_cost is not None and i == 2) else "",
                "",  # level change (manual)
                "",  # reason (manual)
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

    # Fallback if no questions
    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Questions")
        ws.cell(row=1, column=1, value="No questions to export.")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
