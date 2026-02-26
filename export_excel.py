"""
Export generated NEET questions to Excel format.
Groups questions by type, with one sheet per question_type + difficulty.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# Short names for sheet titles
TYPE_SHORT = {
    "MCQ": "MCQ",
    "ASSERTION_REASON": "A_R",
    "MATCH_THE_COLUMN": "MTC",
}


def _format_options(options: dict, correct_answer: str) -> str:
    """Format options as multi-line string with √ marking the correct answer."""
    correct = correct_answer.lower()
    lines = []
    for key in ["a", "b", "c", "d"]:
        if key in options:
            marker = "√ " if key == correct else ""
            lines.append(f"{key}) {marker}{options[key]}")
    return "\n".join(lines)


def export_questions_to_excel(result: dict, time_taken: float = None) -> bytes:
    """
    Export questions to an Excel workbook.

    Each question_type gets its own sheet named {TYPE}_{DIFFICULTY}.
    Column A = question text, Column B = options with √ on correct answer.

    Returns the workbook as bytes (ready for st.download_button).
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    questions = result.get("questions", [])
    metadata = result.get("test_metadata", {})
    difficulty = metadata.get("difficulty", "unknown").upper()

    # Group questions by type
    grouped: dict[str, list] = {}
    for q in questions:
        q_type = q.get("question_type", "MCQ")
        grouped.setdefault(q_type, []).append(q)

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for q_type, q_list in grouped.items():
        short = TYPE_SHORT.get(q_type, q_type.replace(" ", "_").upper())
        sheet_name = f"{short}_{difficulty}"[:31]  # Excel 31-char limit
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = [sheet_name, "OPTIONS"]
        if time_taken is not None:
            headers.append("TIME TAKEN")
        for col, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=value)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 70
        ws.column_dimensions["B"].width = 70
        if time_taken is not None:
            ws.column_dimensions["C"].width = 15

        # Data rows
        for i, q in enumerate(q_list, start=2):
            q_text = q.get("question_text", "")
            options = q.get("options", {})
            correct = q.get("correct_answer", "")

            # Question cell
            cell_q = ws.cell(row=i, column=1, value=q_text)
            cell_q.alignment = cell_alignment
            cell_q.border = thin_border

            # Options cell
            cell_o = ws.cell(row=i, column=2, value=_format_options(options, correct))
            cell_o.alignment = cell_alignment
            cell_o.border = thin_border

            # Time taken cell (only on first data row)
            if time_taken is not None and i == 2:
                cell_t = ws.cell(row=i, column=3, value=f"{time_taken:.1f}s")
                cell_t.alignment = Alignment(horizontal="center", vertical="top")
                cell_t.border = thin_border

    # Fallback if no questions
    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Questions")
        ws.cell(row=1, column=1, value="No questions to export.")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
