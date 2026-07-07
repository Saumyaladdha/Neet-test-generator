"""
Streamlit QA tool — upload ONE PDF or image, generate Hindi questions by
calling the SAME core functions worker/main.py uses in production, directly
and in-process (no API, no SQS, no worker, no answer generation).

For PDFs, this replicates worker/main.py's _process_pdf_job() exactly:
topic detection -> per-topic question distribution -> parallel per-chunk
generation -> dedup. This is what makes a 50-page PDF actually work here —
without it, the whole document would be sent to the model in one call and
hit context_length_exceeded, which production never does because it always
chunks first.

Review each question, mark it, add a comment, then export to Excel.

Run:
    cd test-generator
    source venv/bin/activate
    streamlit run tests/benchmarks/qa_app.py
"""

import base64
import io
import os
import re
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import streamlit as st
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

# On Streamlit Cloud, secrets.toml values are only synced into os.environ
# once st.secrets is accessed — but core.config reads OPENAI_API_KEY via
# os.getenv() at import time, before anything below touches st.secrets.
# Bridge it explicitly so the import chain below doesn't crash first.
if "OPENAI_API_KEY" in st.secrets:
    os.environ.setdefault("OPENAI_API_KEY", st.secrets["OPENAI_API_KEY"])

from core.generator import generate_chunk
from core.pdf import get_page_count_from_bytes, extract_pages
from core.topic_detector import detect_topics, distribute_questions
from core.dedup import dedup_questions

_MAX_PARALLEL_CHUNKS = 4

st.set_page_config(page_title="NEET Hindi Question QA Tool", page_icon="📝", layout="wide")
st.title("📝 NEET Hindi Question Generator — QA Tool")
st.caption("Calls the prompt/LLM directly — no API, SQS, or worker involved. Hindi medium only. No answer generation.")

# ── Session state ───────────────────────────────────────────────────────────

for key, default in [
    ("questions", []), ("source_name", ""), ("marks", {}), ("comments", {}),
]:
    if key not in st.session_state:
        st.session_state[key] = default

QUESTION_TYPE_LABELS = {
    "mcq": "MCQ",
    "assertion_reason": "Assertion-Reason",
    "match_the_column": "Match the Column",
}

# ── LaTeX / tabular cleanup ───────────────────────────────────────────────────

_TABULAR_RE = re.compile(r"\\begin\{tabular\}\{[^}]*\}(.*?)\\end\{tabular\}", re.DOTALL)


def _parse_tabular(tabular_body: str) -> list:
    """Turn a LaTeX tabular body into a list of row-lists of plain strings."""
    body = tabular_body.replace(r"\hline", "")
    rows = [r.strip() for r in body.split(r"\\") if r.strip()]
    table = []
    for row in rows:
        cells = [c.strip() for c in row.split("&")]
        table.append(cells)
    return table


def render_question_text(text: str) -> None:
    """
    Render question text in Streamlit, cleanly:
    - Simple LaTeX math ($...$, $$...$$) is left as-is — Streamlit's
      st.markdown renders it natively via MathJax.
    - LaTeX \\begin{tabular}...\\end{tabular} blocks (used by Match-the-Column
      questions) are NOT math and would show as broken raw text if left in —
      these are extracted and rendered as a real table instead.
    """
    if not text:
        return
    last_end = 0
    for m in _TABULAR_RE.finditer(text):
        before = text[last_end:m.start()].strip()
        if before:
            st.markdown(before)
        table = _parse_tabular(m.group(1))
        if table:
            st.table(table)
        last_end = m.end()
    remainder = text[last_end:].strip()
    if remainder:
        st.markdown(remainder)


def clean_option_text(text: str) -> str:
    """Options are always plain text/simple LaTeX — st.markdown handles them directly."""
    return text if text is not None else ""


# ── Sidebar controls ──────────────────────────────────────────────────────────

with st.sidebar:
    st.header("⚙️ Settings")
    subject = st.selectbox("Subject", ["biology", "chemistry"])
    question_type = st.selectbox(
        "Question Type",
        ["mcq", "assertion_reason", "match_the_column"],
        format_func=lambda x: QUESTION_TYPE_LABELS[x],
    )
    difficulty = st.selectbox("Difficulty", ["easy", "medium", "hard"])
    question_count = st.number_input("Question Count", min_value=1, max_value=50, value=5)
    st.info("Medium: **Hindi** (fixed for this tool)")

_MAX_FILE_MB = 50  # matches the spec's PDF upload limit (Section 02: File Upload Rules)

uploaded = st.file_uploader(
    f"Upload one PDF or image (max {_MAX_FILE_MB} MB, matching production's own limit)",
    type=["pdf", "png", "jpg", "jpeg", "webp"],
)

if uploaded is not None:
    size_mb = len(uploaded.getvalue()) / (1024 * 1024)
    if size_mb > _MAX_FILE_MB:
        st.error(f"'{uploaded.name}' is {size_mb:.1f} MB — over the {_MAX_FILE_MB} MB limit.")
        uploaded = None

generate_clicked = st.button("🚀 Generate Questions", type="primary", disabled=uploaded is None)


def _upload_pdf_bytes(client: OpenAI, pdf_bytes: bytes) -> str:
    """Same pattern as worker/main.py's _openai_upload — write to a real temp
    file with a .pdf suffix so OpenAI's Files API can infer the file type."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    try:
        with open(tmp_path, "rb") as f:
            return client.files.create(file=f, purpose="user_data").id
    finally:
        os.unlink(tmp_path)


def _generate_pdf_with_chunking(client, pdf_bytes, subject, question_type, difficulty, question_count, progress_cb=None):
    """
    Replicates worker/main.py's _process_pdf_job() for a single component:
    detect topics once -> distribute the requested count across topics ->
    generate each topic's chunk in parallel -> merge + dedup.
    """
    total_pages = get_page_count_from_bytes(pdf_bytes)
    if progress_cb:
        progress_cb(f"{total_pages} pages — detecting topics...")

    full_file_id = _upload_pdf_bytes(client, pdf_bytes)
    try:
        detection = detect_topics(full_file_id, total_pages)
    finally:
        try:
            client.files.delete(full_file_id)
        except Exception:
            pass

    topics = detection["topics"]
    counts, skipped = distribute_questions(topics, int(question_count))
    if skipped and progress_cb:
        progress_cb(f"{len(skipped)} topic(s) got 0 questions (more topics than requested questions)")

    work_items = [(topics[i], counts[i]) for i in range(len(topics)) if counts[i] > 0]
    if progress_cb:
        progress_cb(f"{len(topics)} topic(s) identified -> {len(work_items)} chunk(s) to generate")

    def _do_one(topic, count):
        chunk_bytes = extract_pages(pdf_bytes, topic["start_page"], topic["end_page"])
        chunk_file_id = _upload_pdf_bytes(client, chunk_bytes)
        try:
            qs = generate_chunk(
                file_type="pdf",
                openai_file_id=chunk_file_id,
                subject=subject,
                medium="hindi",
                question_type=question_type,
                difficulty=difficulty,
                question_count=count,
            )
            for q in qs:
                q["difficulty"] = difficulty
                q["topic"] = topic.get("topic", "")
                q["subject"] = subject
            return qs
        finally:
            try:
                client.files.delete(chunk_file_id)
            except Exception:
                pass

    all_questions = []
    with ThreadPoolExecutor(max_workers=min(max(len(work_items), 1), _MAX_PARALLEL_CHUNKS)) as pool:
        futures = {pool.submit(_do_one, t, c): t for t, c in work_items}
        done = 0
        for future in as_completed(futures):
            topic = futures[future]
            done += 1
            try:
                all_questions.extend(future.result())
            except Exception as exc:
                if progress_cb:
                    progress_cb(f"Chunk '{topic.get('topic', '?')}' failed: {exc}")
            if progress_cb:
                progress_cb(f"{done}/{len(work_items)} chunk(s) done")

    before = len(all_questions)
    all_questions = dedup_questions("qa-app-session", all_questions)
    if progress_cb and len(all_questions) < before:
        progress_cb(f"Removed {before - len(all_questions)} near-duplicate question(s)")

    return all_questions


# ── Generation ────────────────────────────────────────────────────────────────

if generate_clicked and uploaded is not None:
    file_bytes = uploaded.getvalue()
    is_pdf = uploaded.type == "application/pdf" or uploaded.name.lower().endswith(".pdf")

    status_box = st.empty()

    def _progress(msg: str) -> None:
        status_box.info(msg)

    with st.spinner(f"Generating {question_count} question(s)... this can take a while for larger files."):
        try:
            client = OpenAI()
            if is_pdf:
                questions = _generate_pdf_with_chunking(
                    client, file_bytes, subject, question_type, difficulty,
                    int(question_count), progress_cb=_progress,
                )
            else:
                ext = uploaded.name.rsplit(".", 1)[-1].lower()
                mime = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"
                data_uri = f"data:{mime};base64,{base64.b64encode(file_bytes).decode()}"
                questions = generate_chunk(
                    file_type="image",
                    presigned_url=data_uri,
                    subject=subject,
                    medium="hindi",
                    question_type=question_type,
                    difficulty=difficulty,
                    question_count=int(question_count),
                )

            st.session_state.questions = questions
            st.session_state.source_name = uploaded.name
            st.session_state.marks = {}
            st.session_state.comments = {}
            st.success(f"Generated {len(questions)} question(s) from {uploaded.name}")
        except Exception as exc:
            st.error(f"Generation failed: {exc}")

# ── Results + review ──────────────────────────────────────────────────────────

if st.session_state.questions:
    st.divider()
    st.subheader(f"Generated Questions ({len(st.session_state.questions)}) — {st.session_state.source_name}")

    for i, q in enumerate(st.session_state.questions):
        st.markdown(f"#### Q{i + 1}")
        render_question_text(q.get("question_text", ""))

        opts = q.get("options", {})
        for k in ["a", "b", "c", "d"]:
            if k in opts:
                st.markdown(f"**{k})** {clean_option_text(str(opts[k]))}")

        col1, col2 = st.columns([1, 2])
        with col1:
            mark = st.radio(
                "Mark", ["Correct", "Incorrect", "Partially Correct"],
                key=f"mark_{i}", horizontal=True,
            )
            st.session_state.marks[i] = mark
        with col2:
            comment = st.text_input("Comment (optional)", key=f"comment_{i}")
            st.session_state.comments[i] = comment

        st.divider()

    # ── Excel export ──────────────────────────────────────────────────────────

    def _build_excel() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "QA Review"

        headers = ["PDF Name", "Total Questions", "Q#", "Question", "Option A", "Option B",
                   "Option C", "Option D", "Mark", "Comment"]
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(wrap_text=True, vertical="top")
        border = Border(*(Side(style="thin"),) * 4)

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, border

        widths = [30, 14, 6, 60, 30, 30, 30, 30, 16, 30]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

        questions = st.session_state.questions
        for i, q in enumerate(questions):
            opts = q.get("options", {})
            row = [
                st.session_state.source_name if i == 0 else "",
                len(questions) if i == 0 else "",
                i + 1,
                q.get("question_text", ""),
                opts.get("a", ""), opts.get("b", ""), opts.get("c", ""), opts.get("d", ""),
                st.session_state.marks.get(i, ""),
                st.session_state.comments.get(i, ""),
            ]
            for col, val in enumerate(row, start=2 if False else 1):
                cell = ws.cell(row=i + 2, column=col, value=val)
                cell.alignment, cell.border = cell_align, border

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    st.divider()
    st.download_button(
        "📥 Download Excel (questions + marks + comments)",
        data=_build_excel(),
        file_name=f"qa_review_{Path(st.session_state.source_name).stem}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
