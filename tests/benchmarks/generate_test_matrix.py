"""
Generation test matrix — calls core.generator.generate_chunk() DIRECTLY, in-process.
No API server, no SQS, no worker, no status polling, no answer generation.
Just: upload each PDF once -> call the prompt/LLM function for every
{question_count} x {difficulty} x {question_type} combo -> save questions to Excel.

Question counts are derived per-PDF as [actual_page_count, actual_page_count + 5]
(filenames like "5pagePdf.pdf" do NOT reliably reflect actual page count).

Usage:
    python tests/benchmarks/generate_test_matrix.py --types mcq
    python tests/benchmarks/generate_test_matrix.py --types mcq,assertion_reason,match_the_column

Output:
    tests/benchmarks/results/matrix_<pdf_label>.xlsx
    tests/benchmarks/results/matrix_run_log.json
"""

import argparse
import json
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from core.pdf import get_page_count
from core.generator import generate_chunk

RESULTS_DIR = Path(__file__).parent / "results"

# (path, page_label, medium)
# NOTE: verified 2026-07-06 — 5pagePdf.pdf, 10pagePdf.pdf, 15pagePdf.pdf,
# 20pagePdf.pdf are ALL actually Hindi-language source content (legacy
# non-Unicode font, same family as NCERT-Hindi-Class-12-Biology.pdf) despite
# their filenames suggesting otherwise. medium="hindi" for all 5.
PDFS = [
    ("tests/5pagePdf.pdf", "5pagePdf", "hindi"),
    ("tests/10pagePdf.pdf", "10pagePdf", "hindi"),
    ("tests/15pagePdf.pdf", "15pagePdf", "hindi"),
    ("tests/20pagePdf.pdf", "20pagePdf", "hindi"),
    ("tests/NCERT-Hindi-Class-12-Biology.pdf", "NCERTHindi", "hindi"),
]

DIFFICULTIES = ["easy", "medium", "hard"]
ALL_TYPES = ["mcq", "assertion_reason", "match_the_column"]
TYPE_ABBR = {"mcq": "MCQ", "assertion_reason": "AR", "match_the_column": "MTC"}

HEADERS = ["Question", "Option A", "Option B", "Option C", "Option D",
           "Question Type", "Difficulty", "Elapsed (s)", "Requested", "Generated"]
COL_WIDTHS = [70, 35, 35, 35, 35, 16, 12, 12, 11, 11]


def _upload_pdf(client: OpenAI, path: str) -> str:
    with open(path, "rb") as f:
        uploaded = client.files.create(file=f, purpose="user_data")
    return uploaded.id


def _write_sheet(wb, sheet_name, questions, elapsed, question_type, difficulty, requested):
    ws = wb.create_sheet(title=sheet_name[:31])
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(wrap_text=True, vertical="top")
    border = Border(*(Side(style="thin"),) * 4)

    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, border
    for col, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    if not questions:
        ws.cell(row=2, column=1, value="NO QUESTIONS GENERATED")
        return

    for i, q in enumerate(questions, start=2):
        opts = q.get("options", {})
        row = [
            q.get("question_text", ""),
            opts.get("a", ""), opts.get("b", ""), opts.get("c", ""), opts.get("d", ""),
            question_type, difficulty,
            elapsed if i == 2 else "",
            requested if i == 2 else "",
            len(questions) if i == 2 else "",
        ]
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment, c.border = cell_align, border


_log_lock = threading.Lock()


def _process_pdf(path: str, page_label: str, medium: str, types_to_run: list, raw_log: list) -> None:
    """Runs all combos for ONE PDF, sequentially within itself. Meant to be
    called concurrently (one thread per PDF) by run()."""
    client = OpenAI()  # separate client instance per thread
    real_pages = get_page_count(path)
    question_counts = [real_pages, real_pages + 5]
    print(f"\n{'='*80}\nPDF: {page_label} ({path}) — actual pages={real_pages} "
          f"medium={medium} question_counts={question_counts}\n{'='*80}")

    file_id = _upload_pdf(client, path)
    print(f"  [{page_label}] uploaded -> file_id={file_id}")

    wb = Workbook()
    wb.remove(wb.active)

    for count in question_counts:
        for difficulty in DIFFICULTIES:
            for qtype in types_to_run:
                abbr = TYPE_ABBR[qtype]
                sheet_name = f"{page_label}_{count}q_{difficulty}_{abbr}"
                print(f"  [{page_label}] [{sheet_name}] generating ...")
                t0 = time.time()
                try:
                    questions = generate_chunk(
                        file_type="pdf",
                        openai_file_id=file_id,
                        subject="biology",
                        medium=medium,
                        question_type=qtype,
                        difficulty=difficulty,
                        question_count=count,
                    )
                    status = "OK"
                except Exception as exc:
                    questions = []
                    status = f"ERROR: {exc}"
                elapsed = round(time.time() - t0, 2)
                print(f"  [{page_label}] [{sheet_name}] {status} elapsed={elapsed}s generated={len(questions)}/{count}")

                _write_sheet(wb, sheet_name, questions, elapsed, qtype, difficulty, count)
                with _log_lock:
                    raw_log.append({
                        "pdf": page_label, "sheet": sheet_name, "real_pages": real_pages,
                        "requested": count, "generated": len(questions),
                        "elapsed": elapsed, "status": status,
                    })
                    with open(RESULTS_DIR / "matrix_run_log.json", "w") as f:
                        json.dump(raw_log, f, indent=2, default=str, ensure_ascii=False)

    out_path = RESULTS_DIR / f"matrix_{page_label}.xlsx"
    wb.save(out_path)
    print(f"  [{page_label}] Saved: {out_path}")

    try:
        client.files.delete(file_id)
    except Exception:
        pass


def run(types_to_run: list, pdf_labels: list = None):
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # Preserve log entries for PDFs NOT being re-run in this invocation
    # (e.g. re-running only 4 of 5 PDFs shouldn't wipe the 5th's prior results).
    log_path = RESULTS_DIR / "matrix_run_log.json"
    raw_log = []
    if log_path.exists() and pdf_labels is not None:
        existing = json.load(open(log_path))
        raw_log = [e for e in existing if e["pdf"] not in pdf_labels]

    pdfs_to_run = [p for p in PDFS if pdf_labels is None or p[1] in pdf_labels]
    if not pdfs_to_run:
        print(f"No PDFs matched labels {pdf_labels}. Available: {[p[1] for p in PDFS]}")
        return

    # One PDF per thread — each PDF's own combos still run sequentially
    # within that thread, but all selected PDFs proceed concurrently. I/O-bound
    # (network calls to OpenAI), so threading (not multiprocessing) is
    # appropriate — same pattern as worker/main.py's chunk ThreadPoolExecutor.
    with ThreadPoolExecutor(max_workers=len(pdfs_to_run)) as pool:
        futures = {
            pool.submit(_process_pdf, path, page_label, medium, types_to_run, raw_log): page_label
            for path, page_label, medium in pdfs_to_run
        }
        for future in as_completed(futures):
            page_label = futures[future]
            try:
                future.result()
            except Exception as exc:
                print(f"  [{page_label}] FAILED ENTIRELY: {exc}")

    print(f"\nDone. Raw log: {RESULTS_DIR / 'matrix_run_log.json'}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--types", default="mcq", help="comma-separated: mcq,assertion_reason,match_the_column")
    parser.add_argument("--pdfs", default=None,
                        help="comma-separated page_labels to restrict to, e.g. 5pagePdf,10pagePdf (default: all)")
    args = parser.parse_args()
    run(args.types.split(","), args.pdfs.split(",") if args.pdfs else None)
